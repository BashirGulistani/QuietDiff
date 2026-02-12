from __future__ import annotations

import csv
import os
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook

from .utils import norm_str





@dataclass(frozen=True)
class Table:
    name: str
    columns: list[str]
    rows: list[dict[str, Any]]


def _read_csv(path: str) -> Table:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        cols = [norm_str(c) for c in (reader.fieldnames or [])]
        rows: list[dict[str, Any]] = []
        for r in reader:
            row = {}
            for k, v in r.items():
                kk = norm_str(k)
                row[kk] = v
            rows.append(row)
    return Table(name=os.path.basename(path), columns=cols, rows=rows)



def _read_xlsx(path: str, sheet: str | None) -> Table:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return Table(name=os.path.basename(path), columns=[], rows=[])

    cols = [norm_str(c) for c in header if c is not None]
    col_idx = []
    for i, c in enumerate(header):
        if c is None:
            continue
        if norm_str(c) == "":
            continue
        col_idx.append((i, norm_str(c)))


    rows: list[dict[str, Any]] = []
    for tup in it:
        if tup is None:
            continue
        row: dict[str, Any] = {}
        for i, c in col_idx:
            val = tup[i] if i < len(tup) else None
            row[c] = val
        rows.append(row)
    wb.close()
    return Table(name=os.path.basename(path), columns=cols, rows=rows)


def read_table(path: str, sheet: str | None = None) -> Table:
    p = path.lower()
    if p.endswith(".csv"):
        return _read_csv(path)
    if p.endswith(".xlsx"):
        return _read_xlsx(path, sheet)
    raise ValueError(f"Unsupported file type: {path}")


def ensure_out_dir(out_dir: str) -> str:
    os.makedirs(out_dir, exist_ok=True)
    return out_dir





